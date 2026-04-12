from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from competitor_service import flatten_snapshot_items, get_portfolio_priority_rows, summarize_all_competitors
from lifecycle_store import get_archive_items, get_recent_changes


RUB_FORMAT = '#,##0 "₽"'
RUB_M2_FORMAT = '#,##0.00 "₽/м²"'
AREA_FORMAT = '#,##0.0 "м²"'


def _apply_header_style(ws):
    fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill


def _autowidth(ws):
    for column_cells in ws.columns:
        max_len = 0
        col_idx = column_cells[0].column
        col_letter = get_column_letter(col_idx)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 14), 42)


def _style_sheet(ws):
    _apply_header_style(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autowidth(ws)


def _fill_number_formats(ws, columns: Dict[str, str]):
    header_map = {str(ws.cell(row=1, column=col).value): col for col in range(1, ws.max_column + 1)}
    for header, fmt in columns.items():
        col = header_map.get(header)
        if not col:
            continue
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = fmt


def create_portfolio_report(snapshots: List[Dict], output_path: Optional[str] = None) -> str:
    if output_path is None:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        output_path = str(Path("reports") / f"portfolio_report_{timestamp}.xlsx")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    stats = summarize_all_competitors(snapshots)
    items = flatten_snapshot_items(snapshots)
    changes = get_recent_changes(days=14)
    archive_items = get_archive_items()
    priority_rows = get_portfolio_priority_rows(snapshots)

    summary_rows = [
        {"Показатель": "Конкурентов в базе", "Значение": stats.get("competitors_total", 0)},
        {"Показатель": "Со свободными помещениями", "Значение": stats.get("competitors_with_free", 0)},
        {"Показатель": "Без данных", "Значение": stats.get("competitors_without_data", 0)},
        {"Показатель": "Ошибок загрузки", "Значение": stats.get("competitors_with_errors", 0)},
        {"Показатель": "Найдено помещений", "Значение": stats.get("count", 0)},
        {"Показатель": "Суммарная площадь", "Значение": stats.get("total_area", 0)},
        {"Показатель": "Средневзвешенная цена", "Значение": stats.get("avg_price", 0)},
        {"Показатель": "Суммарная стоимость", "Значение": stats.get("total_price", 0)},
        {"Показатель": "Неподтвержденных объектов", "Значение": stats.get("unconfirmed_count", 0)},
        {"Показатель": "Выбывших объектов в архиве", "Значение": stats.get("removed_count", 0)},
        {"Показатель": "Конкурентов с устаревшими ручными данными", "Значение": stats.get("stale_competitors", 0)},
        {"Показатель": "Конкурентов, которые скоро нужно проверить", "Значение": stats.get("aging_competitors", 0)},
    ]
    df_summary = pd.DataFrame(summary_rows)

    competitor_rows = []
    for snapshot in snapshots:
        competitor = snapshot.get("competitor", {})
        stats_row = snapshot.get("stats", {})
        latest_record = snapshot.get("latest_record") or {}
        freshness = snapshot.get("freshness", {})
        lifecycle = snapshot.get("lifecycle", {})
        competitor_rows.append(
            {
                "Конкурент": competitor.get("name", ""),
                "Режим": "Парсинг сайта" if competitor.get("mode") == "parsed" else "Ручной учет",
                "Свободных помещений": stats_row.get("count", 0),
                "Свободная площадь, м²": stats_row.get("total_area", 0),
                "Средневзвешенная цена, ₽/м²": stats_row.get("avg_price", 0),
                "Суммарная стоимость, ₽": stats_row.get("total_price", 0),
                "Неподтверждено": lifecycle.get("unconfirmed_count", 0),
                "В архиве": lifecycle.get("removed_count", 0),
                "Свежесть данных": freshness.get("freshness_label", ""),
                "Последняя проверка": freshness.get("last_checked_at", ""),
                "Приоритет прозвона": snapshot.get("priority_label", "Низкий"),
                "Причины приоритета": "; ".join(snapshot.get("priority_reasons") or []) or "—",
                "Источник": latest_record.get("source_label", "") or ("Сайт" if competitor.get("mode") == "parsed" else ""),
                "Достоверность": latest_record.get("reliability_label", "") or ("Высокая" if competitor.get("mode") == "parsed" else ""),
                "Комментарий": latest_record.get("comment", ""),
                "Ошибка": snapshot.get("error", ""),
            }
        )
    df_competitors = pd.DataFrame(competitor_rows, columns=["Конкурент", "Режим", "Свободных помещений", "Свободная площадь, м²", "Средневзвешенная цена, ₽/м²", "Суммарная стоимость, ₽", "Неподтверждено", "В архиве", "Свежесть данных", "Последняя проверка", "Приоритет прозвона", "Причины приоритета", "Источник", "Достоверность", "Комментарий", "Ошибка"])

    item_rows = []
    for item in items:
        item_rows.append(
            {
                "Компания": item.get("company", ""),
                "Тип": item.get("type", ""),
                "Название": item.get("title", ""),
                "Площадь, м²": item.get("area"),
                "Ставка за м², ₽": item.get("price_value") or item.get("price_per_sqm"),
                "Общая стоимость, ₽": item.get("total_price_value") or item.get("total_price"),
                "Источник": item.get("source_label", "") or item.get("source_kind", ""),
                "Достоверность": item.get("reliability_label", ""),
                "Ссылка": item.get("source_url") or item.get("url", ""),
            }
        )
    df_items = pd.DataFrame(item_rows, columns=["Компания", "Тип", "Название", "Площадь, м²", "Ставка за м², ₽", "Общая стоимость, ₽", "Источник", "Достоверность", "Ссылка"])

    archive_rows = []
    for row in archive_items:
        archive_rows.append(
            {
                "Компания": row.get("competitor_name", ""),
                "Статус": row.get("status", ""),
                "Тип": row.get("type", ""),
                "Название": row.get("title", ""),
                "Площадь, м²": row.get("area", 0),
                "Ставка за м², ₽": row.get("price_per_sqm", 0),
                "Общая стоимость, ₽": row.get("total_price", 0),
                "Впервые найдено": row.get("first_seen", ""),
                "Последний раз найдено": row.get("last_seen", ""),
                "Ссылка": row.get("source_url", ""),
            }
        )
    df_archive = pd.DataFrame(archive_rows, columns=["Компания", "Статус", "Тип", "Название", "Площадь, м²", "Ставка за м², ₽", "Общая стоимость, ₽", "Впервые найдено", "Последний раз найдено", "Ссылка"])

    change_rows = []
    for row in changes:
        change_rows.append(
            {
                "Дата": row.get("event_at", ""),
                "Компания": row.get("competitor_name", ""),
                "Тип": row.get("type", ""),
                "Название": row.get("title", ""),
                "Событие": row.get("event_type", ""),
                "Старое значение": row.get("old_value", ""),
                "Новое значение": row.get("new_value", ""),
                "Комментарий": row.get("note", ""),
            }
        )
    df_changes = pd.DataFrame(change_rows, columns=["Дата", "Компания", "Тип", "Название", "Событие", "Старое значение", "Новое значение", "Комментарий"])
    df_priority = pd.DataFrame(priority_rows, columns=["competitor_code", "Конкурент", "Приоритет", "Балл", "Причины", "Последняя проверка", "Свежесть", "Неподтверждено", "Свободных помещений", "Свободная площадь, м²"])

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, index=False, sheet_name="Сводка")
        df_competitors.to_excel(writer, index=False, sheet_name="Конкуренты")
        df_items.to_excel(writer, index=False, sheet_name="Помещения")
        df_changes.to_excel(writer, index=False, sheet_name="Изменения")
        df_archive.to_excel(writer, index=False, sheet_name="Архив")
        df_priority.to_excel(writer, index=False, sheet_name="Прозвон")

        for name in ("Сводка", "Конкуренты", "Помещения", "Изменения", "Архив", "Прозвон"):
            ws = writer.book[name]
            _style_sheet(ws)

        _fill_number_formats(writer.book["Сводка"], {"Значение": RUB_FORMAT})
        ws_summary = writer.book["Сводка"]
        for row in range(2, ws_summary.max_row + 1):
            label = ws_summary[f"A{row}"].value
            cell = ws_summary[f"B{row}"]
            if label == "Суммарная площадь":
                cell.number_format = AREA_FORMAT
            elif label == "Средневзвешенная цена":
                cell.number_format = RUB_M2_FORMAT
            elif label == "Суммарная стоимость":
                cell.number_format = RUB_FORMAT

        _fill_number_formats(writer.book["Конкуренты"], {
            "Свободная площадь, м²": AREA_FORMAT,
            "Средневзвешенная цена, ₽/м²": RUB_M2_FORMAT,
            "Суммарная стоимость, ₽": RUB_FORMAT,
        })
        _fill_number_formats(writer.book["Помещения"], {
            "Площадь, м²": AREA_FORMAT,
            "Ставка за м², ₽": RUB_M2_FORMAT,
            "Общая стоимость, ₽": RUB_FORMAT,
        })
        _fill_number_formats(writer.book["Архив"], {
            "Площадь, м²": AREA_FORMAT,
            "Ставка за м², ₽": RUB_M2_FORMAT,
            "Общая стоимость, ₽": RUB_FORMAT,
        })
        _fill_number_formats(writer.book["Прозвон"], {
            "Свободная площадь, м²": AREA_FORMAT,
        })

        for ws_name, header in (("Помещения", "Ссылка"), ("Архив", "Ссылка")):
            ws = writer.book[ws_name]
            header_map = {str(ws.cell(row=1, column=col).value): col for col in range(1, ws.max_column + 1)}
            col = header_map.get(header)
            if col:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        cell.hyperlink = str(cell.value)
                        cell.style = "Hyperlink"

    return output_path
